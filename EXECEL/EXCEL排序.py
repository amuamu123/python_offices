import openpyxl
from openpyxl.styles import Border,Side
import xlwt
import xlsxwriter
import pandas as pd
from pandas import Series,DataFrame

def main():
    today = input('请问今天是DAY几呀?')
    read_xl(xlsxpath,today)
    
def read_xl(xlsxpath,today):
    namelist = []
    numlist = []
    workbook = openpyxl.load_workbook(xlsxpath) #读取目标文件
    booksheet = workbook.active       #获取最后活动表
    rows = booksheet.rows        #获取sheet页的行数据
    columns = booksheet.columns    #获取sheet页的列数据
    i = 1  # 以下代码为迭代所有的行[row是行，column是列]设置为1表示跳过第一行
    for row in rows:  
        i = i + 1
        line = [col.value for col in row]
        name = booksheet.cell(row=i, column=2).value #对第二列的所有行遍历,获取学员姓名
        DAY_before = booksheet.cell(row=i, column=(int(today)+5)).value #对第DAY的列的前两天的所有行遍历
        DAY_now = booksheet.cell(row=i, column=(int(today)+6)).value #对第DAY的列的所有行遍历
        try:
            num = int(DAY_before)-int(DAY_now)
        except:
            name = '【用于检测是否触底】'
            num = 0
        namelist.append(name)
        numlist.append(num)
    workbook.save('排序测试文件.xlsx')
    sort_xl(namelist,numlist)   

def sort_xl(namelist,numlist):
    df = pd.DataFrame({'学员名称':namelist, '闯关数': numlist})
    df.drop([len(df)-1],inplace=True)  #删去检测行，若想检测是否触底，则注销此行
    df['跳跃排名'] = df.闯关数.rank(method='min',ascending=False)
    df.sort_values("跳跃排名",inplace=True)  #到这里已经排好序了，下面是把表格变漂亮
    writer = pd.ExcelWriter('排序后的文件.xlsx',engine='xlsxwriter')#使用我们的xlsxwriter模块
    df.to_excel(writer,sheet_name='sheet1', header = True, index = None)#写入，表名为sheet1，保留表头，不保留索引
    workbook = writer.book #打开表
    worksheet = writer.sheets['sheet1']#打开sheet
    format_border = workbook.add_format({'border':1})   # 设置边框格式
    worksheet.conditional_format('A1:XFD1048576',{'type':'no_blanks', 'format': format_border}) # 这里是核心，根据条件来设置格式
    workbook.close()
if __name__ == '__main__' :
    xlsxpath = r'D:\MyFiles\Forchange\9-随便捣鼓\一堆测试代码\新建 XLSX 工作表.xlsx'
    main()