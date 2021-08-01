import pandas as pd
import openpyxl
import os
import winreg
from pathlib import Path
from openpyxl.styles import PatternFill, Alignment, Side, Border

#数据清洗(path) 
#print(wb.isna()) # 检查是否有缺失值【True 和 NaN 代表的就是缺失数据。】
#print(wb.head())  # 查看前五行【True 和 NaN 代表的就是缺失数据。】
#print(wb.tail())  # 查看前五行【True 和 NaN 代表的就是缺失数据。】
#print(wb[wb.duplicated()])#打印重复行
#print(wb.info()) # 打印整个表数据

## 新建桌面文件夹
Desktoppath = winreg.QueryValueEx(winreg.OpenKey(winreg.HKEY_CURRENT_USER,r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'), "Desktop")[0]#获取电脑系统桌面路径
try:
    os.makedirs(Desktoppath+"\\考场汇总") #创建一个文件夹
except:
    pass
Desktoppath = Desktoppath + '\\考场汇总'

def 数据添加(in_path,out_path):
    # 添加【序号】
    df = pd.read_excel(in_path) # 打开excel表格
    df.sort_values("年名",inplace=True) 
    df['考场排名'] = df.年名.rank(method='first',ascending=True)
    df.sort_values("考场排名",inplace=True)  #到这里已经排好序了，下面是把表格变漂亮
    #调整格式
    #df.sort_values("班级",inplace=True)
    writer = pd.ExcelWriter(out_path,engine='xlsxwriter')#使用我们的xlsxwriter模块
    df.to_excel(writer,sheet_name='sheet1', header = True, index = None)#写入，表名为sheet1，保留表头，不保留索引
    workbook = writer.book #打开表
    worksheet = writer.sheets['sheet1']#打开sheet
    format_border = workbook.add_format({'border':1})   # 设置边框格式
    worksheet.conditional_format('A1:XFD1048576',{'type':'no_blanks', 'format': format_border}) # 这里是核心，根据条件来设置格式
    workbook.close()
    
    wb = openpyxl.load_workbook(out_path)#根据需求写入考场号、座位号
    ws = wb.active
    ws['E1'] = '考场号'
    ws['F1'] = '座位号'
    x = 1 # 11考场专用
    y = 1 # 12考场专用
    for row in ws.iter_rows(min_row=2,min_col=1,max_col=6):
        if row[-3].value <= 80 :
            row[-2].value = '01'
            row[-1].value = '01{:0>2d}'.format(row[-3].value)
        elif row[-3].value <= 440:
            if row[-3].value%40 == 0:
                row[-2].value = '{:0>2d}'.format(row[-3].value//40-1)
                row[-1].value = '{:0>2d}{:0>2d}'.format(row[-3].value//40-1,40)
            else:
                row[-2].value = '{:0>2d}'.format(row[-3].value//40)
                row[-1].value = '{:0>2d}{:0>2d}'.format(row[-3].value//40,row[-3].value%40)
        elif row[-3].value <= 440+(ws.max_row-441)/2:
            row[-2].value = '11'
            row[-1].value = '11{:0>2d}'.format(x)
            x += 1
        else:
            row[-2].value = '12'
            row[-1].value = '12{:0>2d}'.format(y)
            y += 1
    wb.save(out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active# 打开工作表
    ws.delete_cols(3, 2) # 删除多余行
    ws.column_dimensions['A'].width =  21.5#调整列宽
    ws.column_dimensions['B'].width =  21.5
    ws.column_dimensions['C'].width =  21.5
    ws.column_dimensions['D'].width =  21.5
    align = Alignment(horizontal='center', vertical='center')# 定义对齐样式横向居中、纵向居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align# 循环取出单元格，调整表中样式
    wb.save(out_path)

def 打印美化(in_path,out_path):
    wb = openpyxl.load_workbook(in_path)
    ws = wb.active# 打开工作表
    ws.delete_rows(1, 1)# 从第1行，删1行
    ws.delete_cols(3, 1) # 从第3列，删1列
    x = [] 
    for i in range(2,ws.max_row+1):
        if (i-1)//20%2 == 1:
            ws['D'+str(i-20)] = ws['A'+str(i)].value  
            ws['E'+str(i-20)] = ws['B'+str(i)].value  
            ws['F'+str(i-20)] = ws['C'+str(i)].value
            x.append(i)
    n=0
    for i in x:
        i = i-n
        ws.delete_rows(i, 1)# 从第i行，删1行
        n+=1
    ws.column_dimensions['A'].width = 14#调整列宽
    ws.column_dimensions['B'].width = 14#调整列宽
    ws.column_dimensions['C'].width = 14#调整列宽
    ws.column_dimensions['D'].width = 14#调整列宽
    ws.column_dimensions['E'].width = 14#调整列宽
    ws.column_dimensions['F'].width = 14#调整列宽
    for i in range(ws.max_row+1):
        ws.row_dimensions[i].height = 19.4#调整行高
    align = Alignment(horizontal='center', vertical='center')# 定义对齐样式横向居中、纵向居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align# 循环取出单元格，调整表中样式
    wb.save(out_path)

def 数据分组保存(path,name):
    wb = pd.read_excel(path)# 打开excel表格
    wb = wb.drop_duplicates()# 删除重复行
    wb.dropna()#删除缺失值
    grade_df1 = wb.groupby(name) # 按name分组
    n = 1 
    for i in grade_df1:
        writer = Desktoppath+'\\{}班考试表.xlsx'.format(n)
        i[1].to_excel(writer,header = True, index = None) # 生成表格
        n +=1
        wb = openpyxl.load_workbook(writer)
        ws = wb.active# 打开工作表
        ws['E1'] = '考试号'
        for row in ws.iter_rows(min_row=2,min_col=1,max_col=5):
            if len(str(row[-2].value)) == 3:
                row[-1].value = '0'+str(row[-2].value)
            else:
                row[-1].value = str(row[-2].value)
        ws.delete_cols(4, 1) # 从第3列，删1列 
        ws.column_dimensions['A'].width = 14#调整列宽
        ws.column_dimensions['B'].width = 14#调整列宽
        ws.column_dimensions['C'].width = 14#调整列宽
        ws.column_dimensions['D'].width = 14#调整列宽
        align = Alignment(horizontal='center', vertical='center')# 定义对齐样式横向居中、纵向居中
        side = Side('thin') # 定义边样式为细条
        border = Border(top = side,bottom=side, right=side,left=side)# 定义表头边框样式，有底边和右边
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align# 循环取出单元格，调整表中样式
                cell.border = border# 循环取出单元格，调整表中样式
        wb.save(writer)

def main():
    print('请确认表格样式为：姓名、班级、年名\n')
    x = input('输入完整的文件路径，如：C:\八上期中.xls')
    out= Desktoppath + '\\'+'班级表.xlsx'
    out2= Desktoppath + '\\'+'考场表.xlsx'
    数据添加(x,out)# 输入需要操作的excel、输出后的excel
    打印美化(out,out2)
    数据分组保存(out,'班级')
    os.remove(out)

if __name__ == "__main__":
    main()