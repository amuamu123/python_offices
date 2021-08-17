import os
import openpyxl
'''
把要合并的excel放在一个文件夹里，
输入文件夹绝对路径即可
'''
def get_filename(filepath,newname): # 获取文件路径
    global datamall
    datamall=[]
    for files in os.listdir(filepath): 
        files = filepath +'\\'+ files
        read_xl(files)
    new_func(datamall,newname)

def read_xl(xlsxpath):
    workbook = openpyxl.load_workbook(xlsxpath)
    booksheet = workbook.active   
    for j in range(1,booksheet.max_row+1):
        data = []
        for i in range(1,booksheet.max_column+1):
            data.append(booksheet.cell(row= j, column = i).value)
        datamall.append(data)
def new_func(datamall,newname):
    wb=openpyxl.Workbook()#打开工作表
    sheet=wb.active# 获取工作簿的活动表
    for i in datamall :
        sheet.append(i) 
    wb.save(newname)  
def main(filepath,newname): 
    while True:
        folder = os.path.exists(filepath)
        if (not folder):
            print('未找到文件夹，请重新输入')
        else :
            print('开始读取-------\n')
            break
    get_filename(filepath,newname)  #传参
DesktoppathA = ''
DesktoppathB = ''
main(DesktoppathA,'合并后数据A.xlsx')
main(DesktoppathB,'合并后数据B.xlsx')